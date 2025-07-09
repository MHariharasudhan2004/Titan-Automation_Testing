import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Ask user for website name
website_name = input("Enter the website name (e.g., TITAN): ").strip()

# Input/output paths
input_path = "D:/AT/Report/objects2.txt"
output_excel = "generated_testcases_group4.xlsx"

# Actions per object type
test_actions = {
    "Link": ["Verify whether the link is Exist", "Highlight the link", "Click the link"],
    "WebEdit": ["Verify whether the text box is present", "Highlight the text box", "Enter value in the text box"],
    "WebCheckBox": ["Verify whether the checkbox is present", "Highlight the checkbox", "Check the checkbox"],
    "WebRadioGroup": ["Verify whether the radio button group is present", "Highlight the radio button group", "Select a radio option"],
    "WebList": ["Verify whether the dropdown is present", "Highlight the dropdown", "Select value from dropdown"],
    "WebButton": ["Verify whether the button is present", "Highlight the button", "Click the button"],
    "WebElement": ["Verify whether the element is present", "Highlight the element", "Validate the element's property"],
}

# Initial test steps for website launch
rows = [
    ["Website Launch", "URL - Website Launch", "Website Launch", f"Verify whether the {website_name} Website is opened in Google Chrome", "Pending", ""],
    ["Website Launch", "URL - Website Launch", "Website Launch", f"Verify whether the {website_name} Website is already opened in the browser", "Pending", ""],
    ["Website Launch", "URL - Website Launch", "Website Launch", f"Verify whether the {website_name} Page is loaded successfully", "Pending", ""]
]

# Process the object file
with open(input_path, "r", encoding="windows-1252") as f:
    for line in f:
        line = line.strip()
        if ":" not in line or not line:
            continue

        obj_type_full, obj_name = map(str.strip, line.split(":", 1))
        obj_type = obj_type_full.split()[0].strip()

        if obj_name == "(Write error)":
            continue

        actions = test_actions.get(obj_type)
        if actions:
            for action in actions:
                rows.append([
                    obj_type,           
                    obj_name,           
                    action,             
                    f"{action} for '{obj_name}'",  
                    "Pending",          
                ])


columns = ["Main Module", "Sub Module", "Test Case Name", "Test Cases", "Status", "Comments"]
df = pd.DataFrame(rows, columns=columns)
df.to_excel(output_excel, index=False)

wb = load_workbook(output_excel)
ws = wb.active

header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light Blue
status_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light Green

for cell in ws[1]:
    cell.fill = header_fill

for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):  # 'Status' column
    for cell in row:
        if cell.value == "Pending":
            cell.fill = status_fill

wb.save(output_excel)
print(f"\n{len(df)} Test cases saved to '{output_excel}' with grouping by micClass (e.g., WebEdit, Link).")
